import io
import uuid
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy.orm import Session

from app.models.attendance import Attendance
from app.models.department import Department
from app.models.enums import AttendanceStatus, ProjectStatus, TaskPriority, TaskStatus
from app.models.project import Project
from app.models.task import Task
from app.models.user import User
from app.utils.exceptions import ForbiddenError


class ReportService:
    def __init__(self, db: Session):
        self.db = db

    def _style_header(self, ws, headers: list[str]):
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")

        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align

    def _auto_column_width(self, ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val = str(cell.value or "")
                if len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    def export_employees(self, current_user: User, department_id: uuid.UUID | None = None) -> bytes:
        user_role = current_user.role.name if current_user.role else "employee"
        if user_role not in ["admin", "manager"]:
            raise ForbiddenError("You do not have permission to export employee directory reports.")

        query = self.db.query(User).filter(
            User.company_id == current_user.company_id,
            User.is_deleted.is_(False),
        )
        if department_id is not None:
            query = query.filter(User.department_id == department_id)

        employees = query.order_by(User.full_name).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Employees"

        headers = ["Employee ID", "Full Name", "Email", "Department", "Role", "Status", "Joined Date"]
        self._style_header(ws, headers)

        for emp in employees:
            dept_name = emp.department.name if emp.department else "Unassigned"
            role_name = emp.role.name.upper() if emp.role else "EMPLOYEE"
            status = "Active" if emp.is_active else "Inactive"
            joined = emp.created_at.strftime("%Y-%m-%d") if emp.created_at else "—"

            ws.append([
                str(emp.id),
                emp.full_name,
                emp.email,
                dept_name,
                role_name,
                status,
                joined,
            ])

        self._auto_column_width(ws)
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def export_projects(self, current_user: User, status: ProjectStatus | None = None) -> bytes:
        query = self.db.query(Project).filter(
            Project.company_id == current_user.company_id,
            Project.is_deleted.is_(False),
        )
        if status is not None:
            query = query.filter(Project.status == status)

        projects = query.order_by(Project.name).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Projects"

        headers = ["Project ID", "Project Name", "Description", "Status", "Department", "Start Date", "Due Date", "Created Date"]
        self._style_header(ws, headers)

        for proj in projects:
            dept_name = proj.department.name if proj.department else "General"
            ws.append([
                str(proj.id),
                proj.name,
                proj.description or "",
                proj.status.value,
                dept_name,
                proj.start_date.strftime("%Y-%m-%d") if proj.start_date else "—",
                proj.due_date.strftime("%Y-%m-%d") if proj.due_date else "—",
                proj.created_at.strftime("%Y-%m-%d") if proj.created_at else "—",
            ])

        self._auto_column_width(ws)
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def export_tasks(
        self,
        current_user: User,
        project_id: uuid.UUID | None = None,
        status: TaskStatus | None = None,
        priority: TaskPriority | None = None,
    ) -> bytes:
        user_role = current_user.role.name if current_user.role else "employee"
        query = self.db.query(Task).filter(
            Task.company_id == current_user.company_id,
            Task.is_deleted.is_(False),
        )

        if user_role == "employee":
            query = query.filter(Task.assigned_to_id == current_user.id)

        if project_id is not None:
            query = query.filter(Task.project_id == project_id)
        if status is not None:
            query = query.filter(Task.status == status)
        if priority is not None:
            query = query.filter(Task.priority == priority)

        tasks = query.order_by(Task.created_at.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"

        headers = ["Task ID", "Title", "Project", "Assigned To", "Status", "Priority", "Due Date", "Completed At"]
        self._style_header(ws, headers)

        for task in tasks:
            proj_name = task.project.name if task.project else "—"
            assignee = task.assigned_to.full_name if task.assigned_to else "Unassigned"
            ws.append([
                str(task.id),
                task.title,
                proj_name,
                assignee,
                task.status.value,
                task.priority.value,
                task.due_date.strftime("%Y-%m-%d") if task.due_date else "—",
                task.completed_at.strftime("%Y-%m-%d %H:%M UTC") if task.completed_at else "—",
            ])

        self._auto_column_width(ws)
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    def export_attendance(
        self,
        current_user: User,
        employee_id: uuid.UUID | None = None,
        status: AttendanceStatus | None = None,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> bytes:
        user_role = current_user.role.name if current_user.role else "employee"
        query = self.db.query(Attendance).filter(
            Attendance.company_id == current_user.company_id,
        )

        if user_role == "employee":
            query = query.filter(Attendance.employee_id == current_user.id)
        elif employee_id is not None:
            query = query.filter(Attendance.employee_id == employee_id)

        if status is not None:
            query = query.filter(Attendance.status == status)
        if start_date is not None:
            query = query.filter(Attendance.attendance_date >= start_date)
        if end_date is not None:
            query = query.filter(Attendance.attendance_date <= end_date)

        records = query.order_by(Attendance.attendance_date.desc()).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"

        headers = ["Date", "Employee", "Check In", "Check Out", "Working Minutes", "Status", "Punctuality", "Late Minutes"]
        self._style_header(ws, headers)

        for rec in records:
            emp_name = rec.employee.full_name if rec.employee else "Employee"
            check_in_str = rec.check_in.strftime("%H:%M:%S") if rec.check_in else "—"
            check_out_str = rec.check_out.strftime("%H:%M:%S") if rec.check_out else "—"
            punctuality = "Late" if rec.is_late else "On Time"

            ws.append([
                rec.attendance_date.strftime("%Y-%m-%d"),
                emp_name,
                check_in_str,
                check_out_str,
                rec.working_minutes,
                rec.status.value,
                punctuality,
                rec.late_minutes,
            ])

        self._auto_column_width(ws)
        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

import io
import openpyxl
import pytest


def test_export_employees_admin(client, admin_user_token_headers):
    res = client.get("/api/v1/reports/employees/export", headers=admin_user_token_headers)
    assert res.status_code == 200
    assert res.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # Verify readable openpyxl workbook
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert "Employees" in wb.sheetnames
    ws = wb["Employees"]
    assert ws.cell(row=1, column=1).value == "Employee ID"


def test_export_employees_forbidden_for_employee(client, normal_user_token_headers):
    res = client.get("/api/v1/reports/employees/export", headers=normal_user_token_headers)
    assert res.status_code == 403


def test_export_projects(client, admin_user_token_headers):
    res = client.get("/api/v1/reports/projects/export", headers=admin_user_token_headers)
    assert res.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert "Projects" in wb.sheetnames


def test_export_tasks(client, admin_user_token_headers):
    res = client.get("/api/v1/reports/tasks/export", headers=admin_user_token_headers)
    assert res.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert "Tasks" in wb.sheetnames


def test_export_attendance(client, admin_user_token_headers):
    res = client.get("/api/v1/reports/attendance/export", headers=admin_user_token_headers)
    assert res.status_code == 200
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert "Attendance" in wb.sheetnames
